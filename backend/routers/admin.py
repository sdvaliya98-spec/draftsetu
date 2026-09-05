from fastapi import APIRouter, Depends, HTTPException, Query, Response
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import func, cast, Date, or_
from datetime import datetime, timedelta, time
import json
import os
import io
import shutil
import math
from typing import List, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from backend import models, database
from backend.routers.auth import get_admin_user
from backend.routers.documents import get_libreoffice_status

router = APIRouter(prefix="/admin", tags=["admin"])

@router.get("/dashboard-stats")
def get_dashboard_stats(
    db: Session = Depends(database.get_db),
    admin: models.User = Depends(get_admin_user)
):
    """Retrieve system monitoring statistics and usage metrics for the Admin Dashboard."""
    # 1. Total Documents
    total_documents = db.query(models.DocumentSubmission).count()
    
    # 2. Today's Documents
    today_start = datetime.combine(datetime.utcnow().date(), time.min)
    today_documents = db.query(models.DocumentSubmission).filter(
        models.DocumentSubmission.created_at >= today_start
    ).count()

    # 3. Total Users
    total_users = db.query(models.User).count()

    # 4. Active Users (unique user_ids in document submissions)
    active_users = db.query(models.DocumentSubmission.user_id).filter(
        models.DocumentSubmission.user_id != None
    ).distinct().count()

    # 5. Locked Documents
    locked_documents = db.query(models.DocumentSubmission).filter(
        models.DocumentSubmission.is_locked == True
    ).count()

    # 6. Draft Documents
    draft_documents = db.query(models.DocumentSubmission).filter(
        models.DocumentSubmission.is_locked == False
    ).count()

    # 7. Total Templates
    total_templates = db.query(models.DBTemplate).count()

    # 7.5. Total Static Pages
    total_static_pages = db.query(models.StaticPage).count()

    # 8. Today Activity Detail
    today_generated = db.query(models.DocumentSubmission).filter(
        models.DocumentSubmission.created_at >= today_start
    ).count()
    
    today_drafts = db.query(models.DocumentSubmission).filter(
        models.DocumentSubmission.created_at >= today_start,
        models.DocumentSubmission.is_locked == False
    ).count()
    
    today_locked = db.query(models.DocumentSubmission).filter(
        models.DocumentSubmission.updated_at >= today_start,
        models.DocumentSubmission.is_locked == True
    ).count()

    # 9. Template Usage (aggregated in database to avoid in-memory scans)
    templates = db.query(models.DBTemplate.template_id, models.DBTemplate.name).all()
    template_map = {t[0]: t[1] for t in templates}
    # Built-in fallback template names
    template_map["sale_deed_simple"] = "વેચાણ દસ્તાવેજ (Sale Deed)"
    template_map["varasai_pedhinamu"] = "વારસાઈ આંબો / પેઢીનામું (Pedhinamu)"

    if db.bind.dialect.name == 'postgresql':
        template_id_expr = func.jsonb_extract_path_text(models.DocumentSubmission.data_json, 'template_id')
    else:
        template_id_expr = func.json_extract(models.DocumentSubmission.data_json, '$.template_id')

    usage_stats = db.query(
        template_id_expr,
        func.count(models.DocumentSubmission.id)
    ).group_by(template_id_expr).all()

    usage = {}
    for t_id, count in usage_stats:
        if t_id:
            t_name = template_map.get(t_id, t_id)
            usage[t_name] = usage.get(t_name, 0) + count

    # Sort and format template usages
    sorted_usage = sorted(usage.items(), key=lambda x: x[1], reverse=True)
    template_usage = [{"name": name, "value": val} for name, val in sorted_usage]

    # 10. PDF Engine Status
    status_info = get_libreoffice_status()
    if status_info.get("available"):
        engine_name = status_info.get("engine", "")
        if "Word" in engine_name:
            pdf_engine = "Microsoft Word"
        else:
            pdf_engine = "LibreOffice"
    else:
        pdf_engine = "Not Available"

    # 11. Documents Per Day (Last 7 Days)
    days = []
    today = datetime.utcnow().date()
    for i in range(6, -1, -1):
        d = today - timedelta(days=i)
        days.append(d)
        
    day_counts = {d.isoformat(): 0 for d in days}
    start_date = datetime.combine(today - timedelta(days=6), datetime.min.time())
    
    # Aggregated in database to avoid loading all recent rows
    if db.bind.dialect.name == 'postgresql':
        day_expr = cast(models.DocumentSubmission.created_at, Date)
    else:
        day_expr = func.date(models.DocumentSubmission.created_at)

    recent_stats = db.query(
        day_expr.label("day"),
        func.count(models.DocumentSubmission.id).label("count")
    ).filter(
        models.DocumentSubmission.created_at >= start_date
    ).group_by(
        day_expr
    ).all()
    
    for day, count in recent_stats:
        if day:
            day_str = day.isoformat() if hasattr(day, "isoformat") else str(day)
            if day_str in day_counts:
                day_counts[day_str] = count
            
    # Format date for charts: e.g. "23 May"
    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    formatted_per_day = []
    for date_str in sorted(day_counts.keys()):
        dt = datetime.strptime(date_str, "%Y-%m-%d")
        day_label = f"{dt.day} {months[dt.month - 1]}"
        formatted_per_day.append({"date": day_label, "count": day_counts[date_str]})

    # 12. Locked vs Drafts data
    locked_vs_drafts = [
        {"name": "Locked (લોક)", "value": locked_documents},
        {"name": "Drafts (ડ્રાફ્ટ)", "value": draft_documents}
    ]

    return {
        "total_documents": total_documents,
        "today_documents": today_documents,
        "total_users": total_users,
        "active_users": active_users,
        "locked_documents": locked_documents,
        "draft_documents": draft_documents,
        "total_templates": total_templates,
        "total_static_pages": total_static_pages,
        "today_activity": {
            "generated": today_generated,
            "drafts": today_drafts,
            "locked": today_locked
        },
        "template_usage": template_usage,
        "pdf_engine": pdf_engine,
        "documents_per_day": formatted_per_day,
        "locked_vs_drafts": locked_vs_drafts
    }


@router.get("/users")
def get_admin_users(
    db: Session = Depends(database.get_db),
    admin: models.User = Depends(get_admin_user),
    search: str = Query(default="", description="Global filter by username, name, email, mobile, or city"),
    user_search: Optional[str] = Query(default=None, description="Filter specifically by username or full name"),
    role: Optional[str] = Query(default="all", description="Filter by role: all | user | admin"),
    email: Optional[str] = Query(default=None, description="Filter by email"),
    mobile: Optional[str] = Query(default=None, description="Filter by mobile"),
    city: Optional[str] = Query(default=None, description="Filter by city"),
    sort: str = Query(default="newest", description="Sort order: newest | oldest | most_docs | docs_desc | docs_asc | credits_desc | credits_asc | username_asc | username_desc"),
    page: int = Query(default=1, ge=1, description="Page number"),
    page_size: int = Query(default=20, ge=1, le=100, description="Results per page"),
):
    """Read-only paginated user list with non-sensitive user metadata, wallet balance, and doc counts. No N+1 queries."""

    # ── 1. Aggregate subquery: doc count per user (single SQL round-trip) ──────
    doc_counts_sq = (
        db.query(
            models.DocumentSubmission.user_id,
            func.count(models.DocumentSubmission.id).label("doc_count")
        )
        .filter(models.DocumentSubmission.user_id.isnot(None))
        .group_by(models.DocumentSubmission.user_id)
        .subquery()
    )

    # ── 2. Base query: outer-join docs and wallet so all users are included ───
    query = (
        db.query(
            models.User,
            func.coalesce(doc_counts_sq.c.doc_count, 0).label("doc_count"),
            func.coalesce(models.Wallet.current_balance, 0).label("wallet_balance")
        )
        .outerjoin(doc_counts_sq, models.User.id == doc_counts_sq.c.user_id)
        .outerjoin(models.Wallet, models.User.id == models.Wallet.user_id)
    )

    # ── 3. Multi-field Search filter (Global Search) ──────────────────────────
    if search and search.strip():
        term = f"%{search.strip()}%"
        query = query.filter(
            or_(
                models.User.username.ilike(term),
                models.User.full_name.ilike(term),
                models.User.email.ilike(term),
                models.User.mobile_number.ilike(term),
                models.User.city.ilike(term)
            )
        )

    # ── 3.0 Specific User filter (Name / Username) ───────────────────────────
    if user_search and user_search.strip():
        u_term = f"%{user_search.strip()}%"
        query = query.filter(
            or_(
                models.User.username.ilike(u_term),
                models.User.full_name.ilike(u_term)
            )
        )

    # ── 3.1 Role filter ───────────────────────────────────────────────────────
    if role == "user":
        query = query.filter(models.User.is_admin == False)
    elif role == "admin":
        query = query.filter(models.User.is_admin == True)

    # ── 3.2 Specific Column Filters ───────────────────────────────────────────
    if email and email.strip():
        query = query.filter(models.User.email.ilike(f"%{email.strip()}%"))
    if mobile and mobile.strip():
        query = query.filter(models.User.mobile_number.ilike(f"%{mobile.strip()}%"))
    if city and city.strip():
        query = query.filter(models.User.city.ilike(f"%{city.strip()}%"))

    # ── 4. Sort ────────────────────────────────────────────────────────────────
    if sort == "oldest":
        query = query.order_by(models.User.created_at.asc())
    elif sort in ["most_docs", "docs_desc"]:
        query = query.order_by(func.coalesce(doc_counts_sq.c.doc_count, 0).desc(), models.User.id.desc())
    elif sort == "docs_asc":
        query = query.order_by(func.coalesce(doc_counts_sq.c.doc_count, 0).asc(), models.User.id.asc())
    elif sort == "credits_desc":
        query = query.order_by(func.coalesce(models.Wallet.current_balance, 0).desc(), models.User.id.desc())
    elif sort == "credits_asc":
        query = query.order_by(func.coalesce(models.Wallet.current_balance, 0).asc(), models.User.id.asc())
    elif sort in ["username_asc", "username"]:
        query = query.order_by(models.User.username.asc())
    elif sort == "username_desc":
        query = query.order_by(models.User.username.desc())
    else:  # default: newest
        query = query.order_by(models.User.created_at.desc())

    # ── 5. Total count (before pagination) ────────────────────────────────────
    total = query.count()
    total_pages = max(1, math.ceil(total / page_size))
    page = min(page, total_pages)

    # ── 6. Paginate ────────────────────────────────────────────────────────────
    offset = (page - 1) * page_size
    rows = query.offset(offset).limit(page_size).all()

    # ── 7. Serialize safely (NEVER include passwords, hashes, or tokens) ──────
    users_out = []
    for user, doc_count, wallet_balance in rows:
        users_out.append({
            "id": user.id,
            "username": user.username,
            "full_name": user.full_name or "—",
            "email": user.email or "—",
            "mobile_number": user.mobile_number or "—",
            "city": user.city or "—",
            "wallet_balance": int(wallet_balance or 0),
            "auth_provider": user.auth_provider or "local",
            "is_admin": user.is_admin,
            "created_at": user.created_at.isoformat() if user.created_at else None,
            "doc_count": doc_count,
            "status": "Active" if user.is_active else "Disabled",
            "is_active": user.is_active,
            "document_limit": user.document_limit,
        })

    return {
        "users": users_out,
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": total_pages,
    }


from pydantic import BaseModel
import re

class UserStatusUpdate(BaseModel):
    is_active: bool

class AdminUserUpdate(BaseModel):
    full_name: str | None = None
    email: str | None = None
    username: str | None = None
    mobile_number: str | None = None
    city: str | None = None
    is_active: bool | None = None
    document_limit: int | None = None

@router.put("/users/{user_id}")
def update_user_profile(
    user_id: int,
    payload: AdminUserUpdate,
    db: Session = Depends(database.get_db),
    admin: models.User = Depends(get_admin_user)
):
    """
    In-place update of user metadata by Admin.
    Preserves existing user ID, documents, wallet, transactions, and history.
    """
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    fields_set = payload.model_fields_set if hasattr(payload, "model_fields_set") else getattr(payload, "__fields_set__", set())

    # 1. Email validation & case-insensitive uniqueness
    if "email" in fields_set:
        raw_email = (payload.email or "").strip()
        if raw_email:
            cleaned_email = raw_email.lower()
            if not re.match(r"^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$", cleaned_email):
                raise HTTPException(status_code=400, detail="Invalid email format")
            
            # Uniqueness check (case-insensitive) against other users
            dup_email_user = db.query(models.User).filter(
                func.lower(models.User.email) == cleaned_email,
                models.User.id != user.id
            ).first()
            if dup_email_user:
                raise HTTPException(status_code=400, detail="Email is already registered by another user")
            
            user.email = cleaned_email
        else:
            user.email = None

    # 2. Username validation & uniqueness
    if "username" in fields_set:
        raw_username = (payload.username or "").strip()
        if not raw_username:
            raise HTTPException(status_code=400, detail="Username cannot be empty")
        if len(raw_username) < 3:
            raise HTTPException(status_code=400, detail="Username must be at least 3 characters long")
        if not re.match(r"^[a-zA-Z0-9_.-]+$", raw_username):
            raise HTTPException(status_code=400, detail="Username contains invalid characters")
        
        dup_username_user = db.query(models.User).filter(
            func.lower(models.User.username) == raw_username.lower(),
            models.User.id != user.id
        ).first()
        if dup_username_user:
            raise HTTPException(status_code=400, detail="Username is already taken by another user")
        
        user.username = raw_username

    # 3. Mobile validation
    if "mobile_number" in fields_set:
        raw_mobile = (payload.mobile_number or "").strip()
        if raw_mobile:
            digits_only = "".join(c for c in raw_mobile if c.isdigit())
            if len(digits_only) != 10:
                raise HTTPException(status_code=400, detail="Mobile number must be a valid 10-digit number")
            user.mobile_number = digits_only
        else:
            user.mobile_number = None

    # 4. Full Name
    if "full_name" in fields_set:
        raw_name = (payload.full_name or "").strip()
        user.full_name = raw_name if raw_name else None

    # 5. City
    if "city" in fields_set:
        raw_city = (payload.city or "").strip()
        user.city = raw_city if raw_city else None

    # 6. Active status
    if "is_active" in fields_set and payload.is_active is not None:
        if user.id == admin.id and payload.is_active is False:
            raise HTTPException(status_code=400, detail="Admins cannot disable their own accounts")
        user.is_active = payload.is_active

    # 7. Document Limit (Positive int or None for Unlimited)
    if "document_limit" in fields_set:
        if payload.document_limit is not None:
            if payload.document_limit <= 0:
                raise HTTPException(status_code=400, detail="Document limit must be a positive integer or Unlimited")
            user.document_limit = payload.document_limit
        else:
            user.document_limit = None  # Explicit NULL means Unlimited

    db.commit()
    db.refresh(user)

    # Activity Log
    from backend.services.activity_service import log_activity
    log_activity(db, admin.username, "User Profile Updated", "user", user.username)

    return {
        "id": user.id,
        "username": user.username,
        "full_name": user.full_name or "—",
        "email": user.email or "—",
        "mobile_number": user.mobile_number or "—",
        "city": user.city or "—",
        "auth_provider": user.auth_provider or "local",
        "is_admin": user.is_admin,
        "is_active": user.is_active,
        "status": "Active" if user.is_active else "Disabled",
        "document_limit": user.document_limit,
        "created_at": user.created_at.isoformat() if user.created_at else None
    }

@router.put("/users/{user_id}/status")
def update_user_status(
    user_id: int,
    status_data: UserStatusUpdate,
    db: Session = Depends(database.get_db),
    admin: models.User = Depends(get_admin_user)
):
    """Enable or disable a user account. Admins cannot disable themselves."""
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    if user.id == admin.id:
        raise HTTPException(status_code=400, detail="Admins cannot disable their own accounts")

    user.is_active = status_data.is_active
    db.commit()
    db.refresh(user)

    # Log to Activity Logs
    from backend.services.activity_service import log_activity
    action_str = "User Enabled" if status_data.is_active else "User Disabled"
    log_activity(db, admin.username, action_str, "user", user.username)

    return {"id": user.id, "username": user.username, "is_active": user.is_active}



def is_test_account(user: models.User) -> bool:
    """
    Identifies whether an account is a test/demo account based on standard conventions.
    """
    uname = (user.username or "").lower()
    email = (user.email or "").lower()
    full_name = (user.full_name or "").lower()
    test_keywords = ["test", "sample", "temp", "demo", "mock", "fake", "dummy", "pytest"]
    test_email_domains = ["@test.local", "@draftsetu.local", "@example.com", "@test.com", "@localhost"]
    for kw in test_keywords:
        if kw in uname or kw in email or kw in full_name:
            return True
    for domain in test_email_domains:
        if email.endswith(domain):
            return True
    if uname.startswith(("loc_", "goog_", "usr_", "u1", "u2", "adm_", "admin_usr_", "admin_pag_", "admin_test_", "sample_usr_", "sample_bulk_", "reg_usr_", "other_admin_", "test_")):
        return True
    return False


@router.delete("/users/{user_id}")
def delete_user_permanently(
    user_id: int,
    db: Session = Depends(database.get_db),
    admin: models.User = Depends(get_admin_user)
):
    """Permanently delete a user account and associated dependent records. Admin only."""
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    if user.id == admin.id:
        raise HTTPException(status_code=400, detail="Admins cannot delete their own account")

    if user.is_admin and not is_test_account(user):
        raise HTTPException(status_code=400, detail="Cannot delete non-test administrator accounts")

    try:
        username = user.username

        # 1. Clean up Payment Orders
        db.query(models.PaymentOrder).filter(models.PaymentOrder.user_id == user_id).delete(synchronize_session=False)

        # 2. Clean up Wallet Transactions
        db.query(models.WalletTransaction).filter(models.WalletTransaction.user_id == user_id).delete(synchronize_session=False)

        # 3. Clean up Wallet
        db.query(models.Wallet).filter(models.Wallet.user_id == user_id).delete(synchronize_session=False)

        # 4. Clean up Document Submissions and remove generated files from disk
        user_docs = db.query(models.DocumentSubmission).filter(models.DocumentSubmission.user_id == user_id).all()
        for doc in user_docs:
            for fpath in [doc.file_path, doc.final_pdf_path, doc.final_docx_path]:
                if fpath and os.path.exists(fpath):
                    try:
                        os.remove(fpath)
                    except OSError:
                        pass
        db.query(models.DocumentSubmission).filter(models.DocumentSubmission.user_id == user_id).delete(synchronize_session=False)

        # 5. Delete User record
        db.delete(user)
        db.commit()

        # 6. Log Activity
        from backend.services.activity_service import log_activity
        log_activity(db, admin.username, "User Permanently Deleted", "user", username)

        return {"success": True, "message": f"User '{username}' was permanently deleted."}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to permanently delete user: {str(e)}")


class BulkDeleteUsersRequest(BaseModel):
    user_ids: List[int]


@router.post("/users/bulk-delete")
def bulk_delete_users(
    payload: BulkDeleteUsersRequest,
    db: Session = Depends(database.get_db),
    admin: models.User = Depends(get_admin_user)
):
    """
    Bulk delete selected users (e.g. test accounts) and clean up their records safely.
    Current admin is strictly protected. Non-test admin accounts are protected.
    Test admin accounts and test users can be safely bulk-deleted.
    Uses efficient batch database operations within an atomic transaction.
    """
    if not payload.user_ids:
        raise HTTPException(status_code=400, detail="No user IDs provided for deletion")

    if admin.id in payload.user_ids:
        raise HTTPException(status_code=400, detail="Cannot delete your own admin account")

    users = db.query(models.User).filter(models.User.id.in_(payload.user_ids)).all()
    if not users:
        return {"success": True, "deleted_count": 0, "deleted_ids": [], "skipped_ids": []}

    for u in users:
        if u.id == admin.id:
            raise HTTPException(status_code=400, detail="Cannot delete your own admin account")
        if u.is_admin and not is_test_account(u):
            raise HTTPException(status_code=400, detail=f"Cannot delete non-test administrator account '{u.username}'")

    target_ids = [u.id for u in users]
    target_usernames = [u.username for u in users]

    try:
        # 1. Batch delete Payment Orders
        db.query(models.PaymentOrder).filter(models.PaymentOrder.user_id.in_(target_ids)).delete(synchronize_session=False)

        # 2. Batch delete Wallet Transactions
        db.query(models.WalletTransaction).filter(models.WalletTransaction.user_id.in_(target_ids)).delete(synchronize_session=False)

        # 3. Batch delete Wallets
        db.query(models.Wallet).filter(models.Wallet.user_id.in_(target_ids)).delete(synchronize_session=False)

        # 4. Clean up Document files on disk & batch delete Document Submissions
        user_docs = db.query(models.DocumentSubmission).filter(models.DocumentSubmission.user_id.in_(target_ids)).all()
        for doc in user_docs:
            for fpath in [doc.file_path, doc.final_pdf_path, doc.final_docx_path]:
                if fpath and os.path.exists(fpath):
                    try:
                        os.remove(fpath)
                    except OSError:
                        pass
        db.query(models.DocumentSubmission).filter(models.DocumentSubmission.user_id.in_(target_ids)).delete(synchronize_session=False)

        # 5. Batch delete Users
        db.query(models.User).filter(models.User.id.in_(target_ids)).delete(synchronize_session=False)

        db.commit()

        # 6. Log to Activity Logs
        if target_ids:
            from backend.services.activity_service import log_activity
            log_activity(
                db,
                admin.username,
                f"Bulk Deleted {len(target_ids)} users ({', '.join(target_usernames[:5])}{'...' if len(target_usernames) > 5 else ''})",
                "user",
                f"count:{len(target_ids)}"
            )

        return {
            "success": True,
            "deleted_count": len(target_ids),
            "deleted_ids": target_ids,
            "skipped_ids": []
        }
    except Exception as e:
        db.rollback()
        logger.error(f"Error during bulk user deletion: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to bulk delete users: {str(e)}")


@router.get("/users/export-excel")
def export_users_excel(
    db: Session = Depends(database.get_db),
    admin: models.User = Depends(get_admin_user)
):
    """
    Export registered user records to a formatted Excel (.xlsx) spreadsheet.
    Strictly includes only non-sensitive metadata. Admin only.
    """
    doc_counts_sq = (
        db.query(
            models.DocumentSubmission.user_id.label("user_id"),
            func.count(models.DocumentSubmission.id).label("doc_count")
        )
        .filter(models.DocumentSubmission.user_id != None)
        .group_by(models.DocumentSubmission.user_id)
        .subquery()
    )

    query = (
        db.query(
            models.User,
            func.coalesce(doc_counts_sq.c.doc_count, 0).label("doc_count"),
            func.coalesce(models.Wallet.current_balance, 0).label("wallet_balance")
        )
        .outerjoin(doc_counts_sq, models.User.id == doc_counts_sq.c.user_id)
        .outerjoin(models.Wallet, models.User.id == models.Wallet.user_id)
        .order_by(models.User.id.asc())
    )
    rows = query.all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Users"

    headers = [
        "User ID",
        "Full Name",
        "Username",
        "Email",
        "Mobile",
        "City",
        "Credits / Wallet Balance",
        "Documents Count",
        "Auth Provider",
        "Role",
        "Status",
        "Created Date/Time"
    ]

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.append(headers)

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
    ws.row_dimensions[1].height = 26

    for user, doc_count, wallet_balance in rows:
        created_str = user.created_at.strftime("%Y-%m-%d %H:%M:%S") if user.created_at else "—"
        row_data = [
            user.id,
            user.full_name or "—",
            user.username,
            user.email or "—",
            user.mobile_number or "—",
            user.city or "—",
            int(wallet_balance or 0),
            int(doc_count or 0),
            (user.auth_provider or "local").upper(),
            "Admin" if user.is_admin else "User",
            "Active" if user.is_active else "Disabled",
            created_str
        ]
        ws.append(row_data)

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"DraftSetu_Users_Export_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )


@router.get("/payments")
def get_payment_orders(
    search: Optional[str] = Query(None, description="Search by order_id, payment_id, username, email"),
    status: Optional[str] = Query("all", description="all, success, failed, created, fulfillment_pending"),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=10, le=100),
    db: Session = Depends(database.get_db),
    admin: models.User = Depends(get_admin_user)
):
    """
    Admin only: Paginated list and monitoring of Razorpay payment orders.
    Distinguishes Razorpay payment status from DraftSetu wallet fulfillment status.
    """
    query = db.query(models.PaymentOrder).options(
        joinedload(models.PaymentOrder.user),
        joinedload(models.PaymentOrder.wallet_transaction)
    )

    # 1. Status filtering
    if status == "success":
        query = query.filter(models.PaymentOrder.status == "SUCCESS")
    elif status == "failed":
        query = query.filter(models.PaymentOrder.status.in_(["FAILED", "CANCELLED"]))
    elif status == "created":
        query = query.filter(models.PaymentOrder.status == "CREATED")
    elif status == "fulfillment_pending":
        query = query.filter(
            models.PaymentOrder.status == "SUCCESS",
            models.PaymentOrder.wallet_transaction_id == None
        )

    # 2. Search
    if search and search.strip():
        term = f"%{search.strip()}%"
        query = query.join(models.User).filter(
            or_(
                models.PaymentOrder.order_id.ilike(term),
                models.PaymentOrder.payment_id.ilike(term),
                models.User.username.ilike(term),
                models.User.email.ilike(term),
                models.User.full_name.ilike(term)
            )
        )

    # 3. Overall KPI Metrics (unpaginated)
    total_orders = db.query(models.PaymentOrder).count()
    successful_orders = db.query(models.PaymentOrder).filter(models.PaymentOrder.status == "SUCCESS").count()
    failed_orders = db.query(models.PaymentOrder).filter(models.PaymentOrder.status.in_(["FAILED", "CANCELLED"])).count()
    pending_fulfillment = db.query(models.PaymentOrder).filter(
        models.PaymentOrder.status == "SUCCESS",
        models.PaymentOrder.wallet_transaction_id == None
    ).count()

    total_revenue_paise = db.query(func.coalesce(func.sum(models.PaymentOrder.amount), 0)).filter(
        models.PaymentOrder.status == "SUCCESS"
    ).scalar() or 0
    total_revenue_inr = round(total_revenue_paise / 100.0, 2)

    # 4. Pagination
    total = query.count()
    total_pages = max(1, math.ceil(total / page_size)) if total > 0 else 1
    page = min(page, total_pages)
    offset = (page - 1) * page_size

    orders = query.order_by(models.PaymentOrder.created_at.desc()).offset(offset).limit(page_size).all()

    items = []
    for o in orders:
        if o.status == "SUCCESS":
            rzp_status = "PAID"
        elif o.status in ["FAILED", "CANCELLED"]:
            rzp_status = "FAILED"
        else:
            rzp_status = "PENDING"

        if o.status == "SUCCESS" and o.wallet_transaction_id:
            fulfillment_status = "CREDITED"
        elif o.status == "SUCCESS" and not o.wallet_transaction_id:
            fulfillment_status = "FULFILLMENT_PENDING"
        else:
            fulfillment_status = "NOT_APPLICABLE"

        items.append({
            "id": o.id,
            "order_id": o.order_id,
            "payment_id": o.payment_id or "—",
            "plan_id": o.plan_id,
            "amount_inr": round(o.amount / 100.0, 2),
            "credits": o.credits,
            "currency": o.currency,
            "user_id": o.user_id,
            "username": o.user.username if o.user else "Unknown",
            "user_full_name": o.user.full_name if o.user else "—",
            "user_email": o.user.email if o.user else "—",
            "user_mobile": o.user.mobile_number if o.user else "—",
            "raw_status": o.status,
            "razorpay_payment_status": rzp_status,
            "fulfillment_status": fulfillment_status,
            "error_code": o.error_code,
            "error_description": o.error_description,
            "wallet_transaction_id": o.wallet_transaction_id,
            "created_at": o.created_at.isoformat() if o.created_at else None,
            "updated_at": o.updated_at.isoformat() if o.updated_at else None
        })

    return {
        "items": items,
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": total_pages,
        "metrics": {
            "total_orders": total_orders,
            "successful_orders": successful_orders,
            "failed_orders": failed_orders,
            "fulfillment_pending": pending_fulfillment,
            "total_revenue_inr": total_revenue_inr
        }
    }


@router.post("/payments/{order_id}/reconcile")
def reconcile_payment_order(
    order_id: str,
    db: Session = Depends(database.get_db),
    admin: models.User = Depends(get_admin_user)
):
    """
    Admin only: If a payment order is marked SUCCESS but wallet credit was not attached,
    safely fulfills the credits to user's wallet idempotently.
    """
    order = db.query(models.PaymentOrder).filter(models.PaymentOrder.order_id == order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="Payment order not found")

    if order.status != "SUCCESS":
        raise HTTPException(status_code=400, detail="Only successful payments can be fulfilled")

    if order.wallet_transaction_id:
        return {
            "success": True,
            "message": "Payment already fulfilled and credited.",
            "wallet_transaction_id": order.wallet_transaction_id
        }

    from backend.services.wallet_service import WalletService
    wallet = WalletService.adjust_balance(
        db=db,
        user_id=order.user_id,
        credits=order.credits,
        type="CREDIT",
        source="PAYMENT",
        remarks=f"Razorpay Order {order.order_id} (Manual Admin Reconciliation by {admin.username})"
    )

    tx = db.query(models.WalletTransaction).filter(
        models.WalletTransaction.user_id == order.user_id,
        models.WalletTransaction.source == "PAYMENT"
    ).order_by(models.WalletTransaction.created_at.desc()).first()

    if tx:
        order.wallet_transaction_id = tx.id
    db.commit()

    from backend.services.activity_service import log_activity
    log_activity(db, admin.username, f"Reconciled and fulfilled payment order {order.order_id}", "payment", order.order_id)

    return {
        "success": True,
        "message": f"Successfully credited {order.credits} credits to user.",
        "wallet_balance": wallet.current_balance,
        "wallet_transaction_id": order.wallet_transaction_id
    }




@router.get("/activity-logs")
def get_admin_activity_logs(
    db: Session = Depends(database.get_db),
    admin: models.User = Depends(get_admin_user),
    search: str = Query(default="", description="Search by username"),
    action: str = Query(default="", description="Filter by action name"),
    page: int = Query(default=1, ge=1, description="Page number"),
    page_size: int = Query(default=20, ge=1, le=100, description="Page size"),
):
    """Retrieve system activity logs for auditing purposes. Admin only."""
    query = db.query(models.ActivityLog)

    if search and search.strip():
        query = query.filter(models.ActivityLog.username.ilike(f"%{search.strip()}%"))

    if action and action.strip():
        query = query.filter(models.ActivityLog.action == action.strip())

    # Sort newest first
    query = query.order_by(models.ActivityLog.timestamp.desc())

    total = query.count()
    total_pages = max(1, math.ceil(total / page_size))
    page = min(page, total_pages)

    offset = (page - 1) * page_size
    logs = query.offset(offset).limit(page_size).all()

    logs_out = []
    for log in logs:
        logs_out.append({
            "id": log.id,
            "username": log.username,
            "action": log.action,
            "entity_type": log.entity_type,
            "entity_id": log.entity_id,
            "template_name": log.template_name,
            "timestamp": log.timestamp.isoformat() if log.timestamp else None,
        })

    return {
        "logs": logs_out,
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": total_pages,
    }


def get_dir_size(path: str) -> int:
    """Recursively calculate directory size in bytes, ignoring errors."""
    total_size = 0
    if os.path.exists(path):
        if os.path.isdir(path):
            for dirpath, _, filenames in os.walk(path):
                for f in filenames:
                    fp = os.path.join(dirpath, f)
                    if os.path.exists(fp) and not os.path.islink(fp):
                        try:
                            total_size += os.path.getsize(fp)
                        except OSError:
                            pass
        else:
            try:
                total_size = os.path.getsize(path)
            except OSError:
                pass
    return total_size


@router.get("/storage-analytics")
def get_storage_analytics(
    db: Session = Depends(database.get_db),
    admin: models.User = Depends(get_admin_user)
):
    """Retrieve system storage metrics and usage statistics for administrators."""
    from backend.core.config import settings

    # 1. Summary Stats
    total_users = db.query(models.User).count()
    total_documents = db.query(models.DocumentSubmission).count()
    locked_documents = db.query(models.DocumentSubmission).filter(models.DocumentSubmission.is_locked == True).count()
    draft_documents = db.query(models.DocumentSubmission).filter(models.DocumentSubmission.is_locked == False).count()
    total_templates = db.query(models.DBTemplate).count()
    total_static_pages = db.query(models.StaticPage).count()

    # 2. Sizing Metrics
    # Database (including WAL/SHM files)
    db_path = settings.DATABASE_URL.split("///")[-1]
    db_files = [db_path, f"{db_path}-wal", f"{db_path}-shm"]
    db_size = 0
    for f in db_files:
        if os.path.exists(f):
            try:
                db_size += os.path.getsize(f)
            except OSError:
                pass

    # Folders
    uploads_size = get_dir_size(settings.TEMPLATE_STORAGE)
    generated_size = get_dir_size(settings.OUTPUT_DIR)
    temp_size = get_dir_size(settings.TEMP_RENDERS_DIR) + get_dir_size(settings.TEMP_PREVIEWS_DIR)

    # 3. Top 10 Users by Document Count
    doc_counts_sq = (
        db.query(
            models.DocumentSubmission.user_id,
            func.count(models.DocumentSubmission.id).label("doc_count")
        )
        .filter(models.DocumentSubmission.user_id.isnot(None))
        .group_by(models.DocumentSubmission.user_id)
        .subquery()
    )

    top_users_query = (
        db.query(
            models.User.username,
            func.coalesce(doc_counts_sq.c.doc_count, 0).label("doc_count")
        )
        .join(doc_counts_sq, models.User.id == doc_counts_sq.c.user_id)
        .order_by(doc_counts_sq.c.doc_count.desc())
        .limit(10)
        .all()
    )
    top_users = [{"username": row[0], "doc_count": row[1]} for row in top_users_query]

    # 4. Top 10 Templates by Usage Count (aggregated in database to avoid in-memory scans)
    templates = db.query(models.DBTemplate.template_id, models.DBTemplate.name).all()
    template_map = {t[0]: t[1] for t in templates}
    template_map["sale_deed_simple"] = "વેચાણ દસ્તાવેજ (Sale Deed)"
    template_map["varasai_pedhinamu"] = "વારસાઈ આંબો / પેઢીનામું (Pedhinamu)"

    if db.bind.dialect.name == 'postgresql':
        template_id_expr = func.jsonb_extract_path_text(models.DocumentSubmission.data_json, 'template_id')
    else:
        template_id_expr = func.json_extract(models.DocumentSubmission.data_json, '$.template_id')

    usage_stats = db.query(
        template_id_expr,
        func.count(models.DocumentSubmission.id)
    ).group_by(template_id_expr).all()

    usage = {}
    for t_id, count in usage_stats:
        if t_id:
            t_name = template_map.get(t_id, t_id)
            usage[t_name] = usage.get(t_name, 0) + count

    sorted_usage = sorted(usage.items(), key=lambda x: x[1], reverse=True)[:10]
    top_templates = [{"name": name, "count": count} for name, count in sorted_usage]

    return {
        "summary": {
            "total_users": total_users,
            "total_documents": total_documents,
            "locked_documents": locked_documents,
            "draft_documents": draft_documents,
            "total_templates": total_templates,
            "total_static_pages": total_static_pages,
        },
        "storage": {
            "database_bytes": db_size,
            "uploads_bytes": uploads_size,
            "generated_bytes": generated_size,
            "temp_bytes": temp_size,
        },
        "top_users": top_users,
        "top_templates": top_templates,
    }


@router.get("/template-analytics")
def get_template_analytics(
    db: Session = Depends(database.get_db),
    admin: models.User = Depends(get_admin_user)
):
    """Retrieve template usage patterns and audit metrics for administrators."""
    # 1. Base counts
    total_templates = db.query(models.DBTemplate).count()
    total_documents = db.query(models.DocumentSubmission).count()

    # 2. SQL aggregation using JSON extraction (dialect-agnostic)
    if db.bind.dialect.name == 'postgresql':
        template_id_expr = func.jsonb_extract_path_text(models.DocumentSubmission.data_json, 'template_id')
    else:
        template_id_expr = func.json_extract(models.DocumentSubmission.data_json, '$.template_id')

    usage_stats = db.query(
        template_id_expr.label("template_id"),
        func.count(models.DocumentSubmission.id).label("usage_count"),
        func.max(models.DocumentSubmission.created_at).label("last_used")
    ).group_by(template_id_expr).all()

    # 3. Active templates lookup
    db_templates = db.query(models.DBTemplate.template_id, models.DBTemplate.name).filter(models.DBTemplate.is_active == True).all()
    template_map = {t[0]: t[1] for t in db_templates}
    template_map["sale_deed_simple"] = "વેચાણ દસ્તાવેજ (Sale Deed)"
    template_map["varasai_pedhinamu"] = "વારસાઈ આંબો / પેઢીનામું (Pedhinamu)"

    # 4. Compile counts
    usage_dict = {t_id: {"usage_count": 0, "last_used": None} for t_id in template_map.keys()}
    
    for t_id, count, last_used in usage_stats:
        if not t_id:
            continue
        if t_id in usage_dict:
            usage_dict[t_id]["usage_count"] = count
            usage_dict[t_id]["last_used"] = last_used
        else:
            usage_dict[t_id] = {
                "usage_count": count,
                "last_used": last_used
            }

    # 5. Format and convert datetimes to ISO strings
    templates_list = []
    for t_id, info in usage_dict.items():
        templates_list.append({
            "template_id": t_id,
            "template_name": template_map.get(t_id, t_id),
            "usage_count": info["usage_count"],
            "last_used": info["last_used"].isoformat() if info["last_used"] else None
        })

    # 6. Sort by usage_count DESC and take top 20
    templates_list.sort(key=lambda x: x["usage_count"], reverse=True)
    top_20 = templates_list[:20]

    return {
        "total_templates": total_templates,
        "total_documents": total_documents,
        "templates": top_20
    }


@router.get("/template-health")
def get_template_health(
    db: Session = Depends(database.get_db),
    admin: models.User = Depends(get_admin_user)
):
    """Retrieve template health metrics and usage patterns for administrators."""
    from datetime import datetime, timedelta

    # 1. Fetch all templates (active, archived, etc.)
    templates = db.query(models.DBTemplate).all()

    # 2. Extract resolved template ID expressions
    if db.bind.dialect.name == 'postgresql':
        json_template_id = func.jsonb_extract_path_text(models.DocumentSubmission.data_json, 'template_id')
    else:
        json_template_id = func.json_extract(models.DocumentSubmission.data_json, '$.template_id')
        
    resolved_template_id = func.coalesce(
        func.nullif(models.DocumentSubmission.template_id, ''),
        json_template_id
    )

    # 3. Query document generation count and last used date per template
    usage_stats = db.query(
        resolved_template_id.label("template_id"),
        func.count(models.DocumentSubmission.id).label("doc_count"),
        func.max(models.DocumentSubmission.created_at).label("last_used")
    ).group_by(resolved_template_id).all()

    usage_map = {row.template_id: (row.doc_count, row.last_used) for row in usage_stats if row.template_id}

    # 4. Query active users (distinct users who generated documents within last 30 days)
    cutoff = datetime.utcnow() - timedelta(days=30)
    active_users_stats = db.query(
        resolved_template_id.label("template_id"),
        func.count(models.DocumentSubmission.user_id.distinct()).label("active_users")
    ).filter(
        models.DocumentSubmission.created_at >= cutoff
    ).group_by(resolved_template_id).all()

    active_users_map = {row.template_id: row.active_users for row in active_users_stats if row.template_id}

    # 5. Build final response list
    result = []
    for t in templates:
        tpl_id = t.template_id
        doc_count, last_used = usage_map.get(tpl_id, (0, None))
        active_users = active_users_map.get(tpl_id, 0)

        # Get version: prepare for future versioning by displaying "v1" for all current templates
        version_val = getattr(t, "version", "v1") or "v1"

        result.append({
            "template_id": tpl_id,
            "template_name": t.name,
            "status": t.status or ("ACTIVE" if t.is_active else "INACTIVE"),
            "documents_generated": doc_count,
            "last_used": last_used.isoformat() if last_used else None,
            "active_users": active_users,
            "version": version_val
        })

    # Sort by documents_generated DESC (as required by Phase 2/3/4)
    result.sort(key=lambda x: x["documents_generated"], reverse=True)

    return result


@router.get("/documents")
def get_all_documents(
    db: Session = Depends(database.get_db),
    admin: models.User = Depends(get_admin_user),
    search: str = Query(default="", description="Search by username or tracking ID"),
    status: str = Query(default="all", description="Filter by status: all | draft | finalized"),
    page: int = Query(default=1, ge=1, description="Page number"),
    page_size: int = Query(default=20, ge=1, le=100, description="Results per page"),
):
    """Admin-only: paginated list of ALL user documents with owner username and resolved template name."""

    # Base query: join User to get username
    query = (
        db.query(models.DocumentSubmission, models.User.username)
        .outerjoin(models.User, models.DocumentSubmission.user_id == models.User.id)
    )

    # Search: by username or tracking_id
    if search and search.strip():
        term = f"%{search.strip()}%"
        query = query.filter(
            (models.User.username.ilike(term)) |
            (models.DocumentSubmission.tracking_id.ilike(term))
        )

    # Status filter
    if status == "draft":
        query = query.filter(models.DocumentSubmission.is_locked == False)
    elif status == "finalized":
        query = query.filter(models.DocumentSubmission.is_locked == True)

    # Sort newest first
    query = query.order_by(models.DocumentSubmission.created_at.desc())

    # Pagination
    total = query.count()
    total_pages = max(1, math.ceil(total / page_size))
    page = min(page, total_pages)
    offset = (page - 1) * page_size
    rows = query.offset(offset).limit(page_size).all()

    # Build a DBTemplate config map once
    template_config_map: dict = {}
    db_templates = db.query(models.DBTemplate).all()
    for tpl in db_templates:
        template_config_map[tpl.template_id] = tpl

    # Serialize — resolve template_name from DBTemplate map, fall back to raw id
    docs_out = []
    for doc, username in rows:
        # Extract template_id stored in data_json or doc
        raw_template_id = getattr(doc, "template_id", None)
        if not raw_template_id:
            try:
                data = json.loads(doc.data_json or "{}")
                raw_template_id = data.get("template_id")
            except Exception:
                pass

        # Resolve display name: DB map → raw_id fallback → "—"
        tpl = template_config_map.get(raw_template_id)
        if tpl:
            resolved_template_name = tpl.name
        elif raw_template_id:
            # Fallback mapping
            fallback_map = {
                "sale_deed_simple": "વેચાણ દસ્તાવેજ (Sale Deed)",
                "varasai_pedhinamu": "વારસાઈ આંબો / પેઢીનામું (Pedhinamu)"
            }
            resolved_template_name = fallback_map.get(raw_template_id, raw_template_id)
        else:
            resolved_template_name = "—"

        # Parse data_json
        doc_data = {}
        try:
            doc_data = json.loads(doc.data_json or "{}")
        except Exception:
            pass

        # Resolve identity and secondary fields via generic resolvers
        from backend.routers.documents import resolve_document_identity, resolve_document_secondary

        identity_val = resolve_document_identity(doc_data, tpl)
        secondary_val = resolve_document_secondary(doc_data, tpl)

        # has_secondary is True if resolved value is non-empty and not "-"
        has_secondary = secondary_val != "-" and secondary_val != ""

        resolved_doc_identity = identity_val
        resolved_doc_secondary = secondary_val
        resolved_doc_name = identity_val if identity_val != "-" else resolved_template_name

        docs_out.append({
            "id": doc.id,
            "tracking_id": doc.tracking_id,
            "user_id": doc.user_id,
            "username": username or "—",
            "template_id": raw_template_id or "—",
            "template_name": resolved_template_name,
            "document_name": resolved_doc_name,
            "document_identity": resolved_doc_identity,
            "document_secondary": resolved_doc_secondary,
            "has_secondary": has_secondary,
            "is_locked": doc.is_locked,
            "pdf_ready": doc.pdf_ready,
            "pdf_generation_in_progress": doc.pdf_generation_in_progress,
            "buyer_name": doc.buyer_name,
            "survey_no": doc.survey_no,
            "amount": doc.amount,
            "created_at": doc.created_at.isoformat() if doc.created_at else None,
            "updated_at": doc.updated_at.isoformat() if doc.updated_at else None,
            "data_json": doc.data_json,
        })

    return {
        "documents": docs_out,
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": total_pages,
    }


@router.get("/user-details/{user_id}")
def get_user_details(
    user_id: int,
    db: Session = Depends(database.get_db),
    admin: models.User = Depends(get_admin_user),
):
    """Admin-only: full user profile with document statistics. Never exposed to normal users."""
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    # Document statistics
    total_docs = db.query(models.DocumentSubmission).filter(
        models.DocumentSubmission.user_id == user_id
    ).count()
    draft_docs = db.query(models.DocumentSubmission).filter(
        models.DocumentSubmission.user_id == user_id,
        models.DocumentSubmission.is_locked == False
    ).count()
    finalized_docs = db.query(models.DocumentSubmission).filter(
        models.DocumentSubmission.user_id == user_id,
        models.DocumentSubmission.is_locked == True
    ).count()

    return {
        "id": user.id,
        "username": user.username,
        "mobile_number": user.mobile_number or "—",
        "birth_date": user.birth_date or "—",
        "is_admin": user.is_admin,
        "is_active": user.is_active,
        "role": "Admin" if user.is_admin else "User",
        "created_at": user.created_at.isoformat() if user.created_at else None,
        "stats": {
            "total": total_docs,
            "drafts": draft_docs,
            "finalized": finalized_docs,
        },
    }


@router.get("/template-analytics/{template_id}")
def get_template_analytics_detail(
    template_id: str,
    db: Session = Depends(database.get_db),
    admin: models.User = Depends(get_admin_user)
):
    """Retrieve detailed analytics for a specific template. Admin only."""
    # 1. Fetch template details (can be archived too, so no is_active constraint)
    db_template = db.query(models.DBTemplate).filter(
        models.DBTemplate.template_id == template_id
    ).first()
    
    if db_template:
        template_name = db_template.name
    else:
        # Fallback names
        fallback_map = {
            "sale_deed_simple": "વેચાણ દસ્તાવેજ (Sale Deed)",
            "varasai_pedhinamu": "વારસાઈ આંબો / પેઢીનામું (Pedhinamu)"
        }
        if template_id in fallback_map:
            template_name = fallback_map[template_id]
        else:
            raise HTTPException(status_code=404, detail=f"Template '{template_id}' not found")

    # 2. Extract resolved template ID expressions (dialect-agnostic)
    if db.bind.dialect.name == 'postgresql':
        json_template_id = func.jsonb_extract_path_text(models.DocumentSubmission.data_json, 'template_id')
    else:
        json_template_id = func.json_extract(models.DocumentSubmission.data_json, '$.template_id')
        
    resolved_template_id = func.coalesce(
        func.nullif(models.DocumentSubmission.template_id, ''),
        json_template_id
    )

    # 3. Base Stats
    # Documents Generated
    documents_generated = db.query(models.DocumentSubmission).filter(
        resolved_template_id == template_id
    ).count()

    # Draft Count
    draft_count = db.query(models.DocumentSubmission).filter(
        resolved_template_id == template_id,
        models.DocumentSubmission.is_locked == False
    ).count()

    # Finalized Count
    finalized_count = db.query(models.DocumentSubmission).filter(
        resolved_template_id == template_id,
        models.DocumentSubmission.is_locked == True
    ).count()

    # Active Users in last 30 days
    cutoff_30d = datetime.utcnow() - timedelta(days=30)
    active_users = db.query(
        func.count(models.DocumentSubmission.user_id.distinct())
    ).filter(
        resolved_template_id == template_id,
        models.DocumentSubmission.created_at >= cutoff_30d
    ).scalar() or 0

    # Last Generated Date
    last_gen_dt = db.query(func.max(models.DocumentSubmission.created_at)).filter(
        resolved_template_id == template_id,
        models.DocumentSubmission.is_locked == True
    ).scalar()
    last_generated = last_gen_dt.isoformat() if last_gen_dt else None

    # Last Used Date
    last_used_dt = db.query(func.max(models.DocumentSubmission.created_at)).filter(
        resolved_template_id == template_id
    ).scalar()
    last_used = last_used_dt.isoformat() if last_used_dt else None

    # 4. Top Users (limit 10)
    top_users_query = db.query(
        models.User.username,
        func.count(models.DocumentSubmission.id).label("count"),
        func.max(models.DocumentSubmission.created_at).label("last_active")
    ).join(
        models.DocumentSubmission,
        models.User.id == models.DocumentSubmission.user_id
    ).filter(
        resolved_template_id == template_id
    ).group_by(
        models.User.username
    ).order_by(
        func.count(models.DocumentSubmission.id).desc()
    ).limit(10).all()

    top_users = [
        {
            "username": row[0],
            "count": row[1],
            "last_active": row[2].isoformat() if row[2] else None
        }
        for row in top_users_query
    ]

    # 5. Monthly Trend (last 12 months)
    # Define month string expression
    if db.bind.dialect.name == 'postgresql':
        month_expr = func.to_char(models.DocumentSubmission.created_at, 'YYYY-MM')
    else:
        month_expr = func.strftime('%Y-%m', models.DocumentSubmission.created_at)

    today = datetime.utcnow().date()
    months_list = []
    for i in range(11, -1, -1):
        y = today.year
        m = today.month - i
        while m <= 0:
            m += 12
            y -= 1
        months_list.append(f"{y:04d}-{m:02d}")

    # Start date is the 1st of the month 11 months ago
    start_year = today.year
    start_month = today.month - 11
    while start_month <= 0:
        start_month += 12
        start_year -= 1
    start_date = datetime(start_year, start_month, 1)

    trend_stats = db.query(
        month_expr.label("month"),
        func.count(models.DocumentSubmission.id).label("count")
    ).filter(
        resolved_template_id == template_id,
        models.DocumentSubmission.created_at >= start_date
    ).group_by(
        month_expr
    ).all()

    trend_map = {row.month: row.count for row in trend_stats if row.month}
    monthly_trend = [{"month": m, "count": trend_map.get(m, 0)} for m in months_list]

    return {
        "template_id": template_id,
        "template_name": template_name,
        "documents_generated": documents_generated,
        "draft_count": draft_count,
        "finalized_count": finalized_count,
        "active_users": active_users,
        "last_generated": last_generated,
        "last_used": last_used,
        "top_users": top_users,
        "monthly_trend": monthly_trend
    }


@router.get("/template-analytics/{template_id}/documents")
def get_template_analytics_documents(
    template_id: str,
    status: str = Query("all", pattern="^(all|draft|finalized)$"),
    db: Session = Depends(database.get_db),
    admin: models.User = Depends(get_admin_user)
):
    """Retrieve all document submissions for a specific template. Admin only."""
    if db.bind.dialect.name == 'postgresql':
        json_template_id = func.jsonb_extract_path_text(models.DocumentSubmission.data_json, 'template_id')
    else:
        json_template_id = func.json_extract(models.DocumentSubmission.data_json, '$.template_id')
        
    resolved_template_id = func.coalesce(
        func.nullif(models.DocumentSubmission.template_id, ''),
        json_template_id
    )

    query = db.query(
        models.DocumentSubmission,
        models.User.username
    ).outerjoin(
        models.User,
        models.User.id == models.DocumentSubmission.user_id
    ).filter(
        resolved_template_id == template_id
    )

    if status == "draft":
        query = query.filter(models.DocumentSubmission.is_locked == False)
    elif status == "finalized":
        query = query.filter(models.DocumentSubmission.is_locked == True)

    submissions = query.order_by(models.DocumentSubmission.created_at.desc()).all()

    results = []
    for sub, username in submissions:
        results.append({
            "tracking_id": sub.tracking_id,
            "status": "Finalized" if sub.is_locked else "Draft",
            "is_locked": sub.is_locked,
            "created_at": sub.created_at.isoformat() if sub.created_at else None,
            "user_id": sub.user_id,
            "username": username or "Unknown",
            "data_json": sub.data_json
        })
    return results


@router.get("/template-analytics/{template_id}/users")
def get_template_analytics_users(
    template_id: str,
    db: Session = Depends(database.get_db),
    admin: models.User = Depends(get_admin_user)
):
    """Retrieve user-wise aggregation of document submissions for a specific template. Admin only."""
    from sqlalchemy import case

    if db.bind.dialect.name == 'postgresql':
        json_template_id = func.jsonb_extract_path_text(models.DocumentSubmission.data_json, 'template_id')
    else:
        json_template_id = func.json_extract(models.DocumentSubmission.data_json, '$.template_id')
        
    resolved_template_id = func.coalesce(
        func.nullif(models.DocumentSubmission.template_id, ''),
        json_template_id
    )

    query = db.query(
        models.DocumentSubmission.user_id,
        models.User.username,
        func.count(models.DocumentSubmission.id).label("documents"),
        func.sum(case((models.DocumentSubmission.is_locked == False, 1), else_=0)).label("drafts"),
        func.sum(case((models.DocumentSubmission.is_locked == True, 1), else_=0)).label("finalized"),
        func.max(models.DocumentSubmission.created_at).label("last_activity")
    ).join(
        models.User,
        models.User.id == models.DocumentSubmission.user_id
    ).filter(
        resolved_template_id == template_id
    ).group_by(
        models.DocumentSubmission.user_id,
        models.User.username
    ).order_by(
        func.count(models.DocumentSubmission.id).desc()
    )

    rows = query.all()
    results = []
    for row in rows:
        results.append({
            "user_id": row.user_id,
            "username": row.username,
            "documents": row.documents,
            "drafts": int(row.drafts or 0),
            "finalized": int(row.finalized or 0),
            "last_activity": row.last_activity.isoformat() if row.last_activity else None
        })
    return results

